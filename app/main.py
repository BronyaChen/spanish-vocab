"""
FastAPI 主应用
定义所有 API 路由，挂载静态文件。
"""
import csv
import io
import json
import random
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List

import httpx
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db, get_setting, set_setting, Word, Setting
from app.ocr import extract_words, PROVIDERS, _parse_json_response

# ============================================================
# FastAPI 实例
# ============================================================
app = FastAPI(title="西班牙语背单词 App", version="1.0.0")

# 挂载静态文件目录
static_dir = Path(__file__).parent.parent / "static"
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")


# ============================================================
# 根路径 → 返回前端页面
# ============================================================
@app.get("/")
async def index():
    index_file = static_dir / "index.html"
    if index_file.exists():
        return FileResponse(str(index_file))
    return JSONResponse({"message": "前端页面尚未部署，请将 index.html 放入 static/ 目录"})


# ============================================================
# /sw.js — 将 Service Worker serve 到根路径
# 以根路径提供并允许 scope=/，使 SW 能控制整个站点
# ============================================================
@app.get("/sw.js")
async def service_worker():
    return FileResponse(
        str(static_dir / "sw.js"),
        media_type="application/javascript",
        headers={"Service-Worker-Allowed": "/"},
    )


# ============================================================
# Pydantic 模型
# ============================================================
class ActionBody(BaseModel):
    action: str  # remember / favorite / kill

class BatchWordsBody(BaseModel):
    words: List[dict]

class SettingsBody(BaseModel):
    ai_provider: Optional[str] = None
    qwen_api_key: Optional[str] = None
    openai_api_key: Optional[str] = None
    quiz_days: Optional[int] = None
    quiz_limit: Optional[int] = None

class ImportBody(BaseModel):
    words: List[dict]


# ============================================================
# GET /api/quiz — 获取测验单词（加权随机）
# ============================================================
@app.get("/api/quiz")
def api_quiz(
    mode: str = Query("es2cn", regex="^(es2cn|cn2es)$"),
    days: Optional[int] = None,
    limit: Optional[int] = None,
    db: Session = Depends(get_db),
):
    # 从设置读取默认值
    if days is None:
        days = int(get_setting(db, "quiz_days") or "3")
    if limit is None:
        limit = int(get_setting(db, "quiz_limit") or "20")

    # 截止时间：now - days 天
    cutoff = datetime.utcnow() - timedelta(days=days)

    # 取出所有单词，按权重分组
    all_words = db.query(Word).all()
    if not all_words:
        return {"words": []}

    # 权重分组
    group4 = []  # 从未复习过
    group3 = []  # 超过 X 天未复习（非 killed）
    group2 = []  # 收藏夹
    group1 = []  # 已斩

    for w in all_words:
        if w.last_reviewed_at is None and w.status != "killed":
            group4.append(w)
        elif w.status == "killed":
            group1.append(w)
        elif w.status == "favorite":
            group2.append(w)
        elif w.last_reviewed_at is not None and w.last_reviewed_at < cutoff:
            group3.append(w)
        # 最近复习过且不属于以上分类的单词不参与本轮抽查

    # 构造加权池
    weighted_pool = []  # (word, weight)
    for w in group4:
        weighted_pool.append((w, 4))
    for w in group3:
        weighted_pool.append((w, 3))
    for w in group2:
        weighted_pool.append((w, 2))
    for w in group1:
        weighted_pool.append((w, 1))

    if not weighted_pool:
        return {"words": []}

    # 加权采样（不重复）
    actual_limit = min(limit, len(weighted_pool))
    words_list = [item[0] for item in weighted_pool]
    weights = [item[1] for item in weighted_pool]

    selected = []
    selected_ids = set()
    # 逐个采样避免重复
    remaining_indices = list(range(len(weighted_pool)))
    remaining_weights = weights[:]

    for _ in range(actual_limit):
        if not remaining_indices:
            break
        chosen = random.choices(remaining_indices, weights=remaining_weights, k=1)[0]
        idx_in_remaining = remaining_indices.index(chosen)
        selected.append(words_list[chosen])
        selected_ids.add(words_list[chosen].id)
        remaining_indices.pop(idx_in_remaining)
        remaining_weights.pop(idx_in_remaining)

    return {
        "words": [
            {"id": w.id, "spanish": w.spanish, "english": w.english, "chinese": w.chinese}
            for w in selected
        ]
    }


# ============================================================
# POST /api/words/{id}/action — 单词操作
# ============================================================
@app.post("/api/words/{word_id}/action")
def api_word_action(word_id: int, body: ActionBody, db: Session = Depends(get_db)):
    word = db.query(Word).filter(Word.id == word_id).first()
    if not word:
        raise HTTPException(status_code=404, detail="单词不存在")

    if body.action == "remember":
        word.last_reviewed_at = datetime.utcnow()
    elif body.action == "favorite":
        word.status = "favorite"
        word.last_reviewed_at = datetime.utcnow()
    elif body.action == "kill":
        word.status = "killed"
    else:
        raise HTTPException(status_code=400, detail=f"无效操作：{body.action}")

    db.commit()
    return {"ok": True}


# ============================================================
# POST /api/words/{id}/restore — 恢复已删除单词
# ============================================================
@app.post("/api/words/{word_id}/restore")
def api_word_restore(word_id: int, db: Session = Depends(get_db)):
    word = db.query(Word).filter(Word.id == word_id).first()
    if not word:
        raise HTTPException(status_code=404, detail="单词不存在")

    word.status = "active"
    db.commit()
    return {"ok": True}


# ============================================================
# GET /api/words — 单词列表（分页）
# ============================================================
@app.get("/api/words")
def api_words(
    view: str = Query("all", regex="^(all|favorites|killed)$"),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    query = db.query(Word)

    if view == "favorites":
        query = query.filter(Word.status == "favorite")
    elif view == "killed":
        query = query.filter(Word.status == "killed")

    total = query.count()
    words = query.offset((page - 1) * size).limit(size).all()

    return {
        "words": [
            {"id": w.id, "spanish": w.spanish, "english": w.english, "chinese": w.chinese, "status": w.status}
            for w in words
        ],
        "total": total,
        "page": page,
        "size": size,
    }


# ============================================================
# POST /api/upload/excel — 上传 Excel/CSV 解析单词
# ============================================================
@app.post("/api/upload/excel")
async def api_upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="未提供文件")

    filename_lower = file.filename.lower()
    if not any(filename_lower.endswith(ext) for ext in (".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls / .csv 格式的文件")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="文件内容为空")

    words = []
    parse_error = None

    try:
        if filename_lower.endswith(".csv"):
            words = _parse_csv(file_bytes)
        else:
            words = _parse_excel(file_bytes)
    except Exception as e:
        parse_error = str(e)

    # 标准解析成功
    if words and not parse_error:
        return {"words": words}

    # 标准解析失败，尝试 AI 解析
    provider = get_setting(db, "ai_provider") or "qwen"
    api_key = get_setting(db, f"{provider}_api_key") or ""

    if not api_key:
        # 没有 AI key，直接报错
        detail = parse_error or "文件解析失败"
        raise HTTPException(status_code=400, detail=f"文件标准解析失败（{detail}），且未配置 AI Key 无法进行智能解析")

    # 用 AI 解析文件内容
    try:
        ai_words = await _ai_parse_file_content(provider, api_key, file_bytes, filename_lower)
        if ai_words:
            return {"words": ai_words}
        raise HTTPException(status_code=400, detail="AI 未能从文件中识别出有效单词数据")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"AI 解析失败：{str(e)}")


def _parse_csv(file_bytes: bytes) -> List[dict]:
    """解析 CSV 文件：第一列西班牙语、第二列英文、第三列中文，跳过标题行"""
    # 尝试不同编码
    text = None
    for encoding in ("utf-8", "utf-8-sig", "gbk", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except (UnicodeDecodeError, ValueError):
            continue

    if text is None:
        raise ValueError("无法识别文件编码")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if len(rows) < 2:
        raise ValueError("文件行数不足（至少需要标题行+1行数据）")

    words = []
    for row in rows[1:]:  # 跳过标题行
        if len(row) < 2:
            continue
        word = {
            "spanish": row[0].strip(),
            "english": row[1].strip() if len(row) > 1 else "",
            "chinese": row[2].strip() if len(row) > 2 else "",
        }
        if word["spanish"]:
            words.append(word)

    if not words:
        raise ValueError("未能从 CSV 中解析出有效单词")
    return words


def _parse_excel(file_bytes: bytes) -> List[dict]:
    """解析 Excel 文件：第一列西班牙语、第二列英文、第三列中文，跳过标题行"""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("文件行数不足（至少需要标题行+1行数据）")

    words = []
    for row in rows[1:]:  # 跳过标题行
        if len(row) < 2:
            continue
        spanish = str(row[0]).strip() if row[0] else ""
        english = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        chinese = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if spanish:
            words.append({"spanish": spanish, "english": english, "chinese": chinese})

    if not words:
        raise ValueError("未能从 Excel 中解析出有效单词")
    return words


async def _ai_parse_file_content(provider: str, api_key: str, file_bytes: bytes, filename: str) -> List[dict]:
    """调用 AI 解析文件内容为结构化单词数据"""
    # 将文件内容转为文本以发送给 AI
    text_content = None
    if filename.endswith(".csv"):
        for encoding in ("utf-8", "utf-8-sig", "gbk", "latin-1"):
            try:
                text_content = file_bytes.decode(encoding)
                break
            except (UnicodeDecodeError, ValueError):
                continue
    else:
        # Excel: 尝试读取所有单元格为文本
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
            ws = wb.active
            lines = []
            for row in ws.iter_rows(values_only=True):
                lines.append("\t".join(str(c) if c else "" for c in row))
            text_content = "\n".join(lines)
        except Exception:
            pass

    if not text_content:
        raise ValueError("无法读取文件内容")

    # 构造 AI 请求
    config = PROVIDERS[provider]
    prompt = (
        "以下是一个单词表文件的内容，请从中提取所有西班牙语单词及其翻译。"
        "返回严格的 JSON 数组，格式为：\n"
        '[{"spanish":"单词","english":"英文翻译","chinese":"中文翻译"}]\n'
        "要求：\n"
        "1. spanish 是西班牙语单词/词组\n"
        "2. english 是英文翻译\n"
        "3. chinese 是中文翻译\n"
        "4. 只返回 JSON，不要任何其他说明文字\n\n"
        f"文件内容：\n{text_content[:8000]}"
    )

    payload = {
        "model": config["model"].replace("qwen-vl-max", "qwen-max"),  # 文本任务用文本模型
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": 4096,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post(config["base_url"], json=payload, headers=headers)
        resp.raise_for_status()

    result = resp.json()
    content = result["choices"][0]["message"]["content"]
    return _parse_json_response(content)


# ============================================================
# POST /api/upload — 上传截图 OCR 识别
# ============================================================
@app.post("/api/upload")
async def api_upload(file: UploadFile = File(...), db: Session = Depends(get_db)):
    # 读取设置
    provider = get_setting(db, "ai_provider") or "qwen"
    api_key = get_setting(db, f"{provider}_api_key") or ""

    if not api_key:
        raise HTTPException(status_code=400, detail=f"请先在设置中配置 {provider.upper()} 的 API Key")

    # 读取图片
    image_bytes = await file.read()

    try:
        words = await extract_words(provider, api_key, image_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {"words": words}


# ============================================================
# POST /api/words/batch — 批量入库
# ============================================================
@app.post("/api/words/batch")
def api_words_batch(body: BatchWordsBody, db: Session = Depends(get_db)):
    inserted = 0
    skipped = 0

    for item in body.words:
        spanish = item.get("spanish", "").strip()
        if not spanish:
            skipped += 1
            continue

        existing = db.query(Word).filter(Word.spanish == spanish).first()
        if existing:
            # 已存在：更新 english/chinese（如果提供了新值）
            updated = False
            if item.get("english") and item["english"].strip():
                existing.english = item["english"].strip()
                updated = True
            if item.get("chinese") and item["chinese"].strip():
                existing.chinese = item["chinese"].strip()
                updated = True
            if updated:
                db.commit()
            skipped += 1
        else:
            word = Word(
                spanish=spanish,
                english=item.get("english", "").strip() or None,
                chinese=item.get("chinese", "").strip() or None,
            )
            db.add(word)
            inserted += 1

    db.commit()
    return {"inserted": inserted, "skipped": skipped}


# ============================================================
# GET /api/settings — 获取设置
# ============================================================
@app.get("/api/settings")
def api_get_settings(db: Session = Depends(get_db)):
    return {
        "ai_provider": get_setting(db, "ai_provider") or "qwen",
        "qwen_api_key": get_setting(db, "qwen_api_key") or "",
        "openai_api_key": get_setting(db, "openai_api_key") or "",
        "quiz_days": int(get_setting(db, "quiz_days") or "3"),
        "quiz_limit": int(get_setting(db, "quiz_limit") or "20"),
    }


# ============================================================
# PUT /api/settings — 更新设置
# ============================================================
@app.put("/api/settings")
def api_put_settings(body: SettingsBody, db: Session = Depends(get_db)):
    if body.ai_provider is not None:
        set_setting(db, "ai_provider", body.ai_provider)
    if body.qwen_api_key is not None:
        set_setting(db, "qwen_api_key", body.qwen_api_key)
    if body.openai_api_key is not None:
        set_setting(db, "openai_api_key", body.openai_api_key)
    if body.quiz_days is not None:
        set_setting(db, "quiz_days", str(body.quiz_days))
    if body.quiz_limit is not None:
        set_setting(db, "quiz_limit", str(body.quiz_limit))
    return {"ok": True}


# ============================================================
# GET /api/backup/export — 导出所有单词
# ============================================================
@app.get("/api/backup/export")
def api_backup_export(db: Session = Depends(get_db)):
    words = db.query(Word).all()
    return JSONResponse(
        content=[
            {
                "spanish": w.spanish,
                "english": w.english,
                "chinese": w.chinese,
                "status": w.status,
                "created_at": w.created_at.isoformat() if w.created_at else None,
                "last_reviewed_at": w.last_reviewed_at.isoformat() if w.last_reviewed_at else None,
            }
            for w in words
        ],
        headers={"Content-Disposition": "attachment; filename=words_backup.json"},
    )


# ============================================================
# POST /api/backup/import — 导入单词
# ============================================================
@app.post("/api/backup/import")
def api_backup_import(body: ImportBody, db: Session = Depends(get_db)):
    inserted = 0
    skipped = 0

    for item in body.words:
        spanish = item.get("spanish", "").strip()
        if not spanish:
            skipped += 1
            continue

        existing = db.query(Word).filter(Word.spanish == spanish).first()
        if existing:
            skipped += 1
        else:
            word = Word(
                spanish=spanish,
                english=item.get("english", "").strip() or None,
                chinese=item.get("chinese", "").strip() or None,
                status=item.get("status", "active"),
            )
            db.add(word)
            inserted += 1

    db.commit()
    return {"inserted": inserted, "skipped": skipped}
